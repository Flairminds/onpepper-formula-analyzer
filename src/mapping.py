
import logging

# Configure a simple logger for this module
logger = logging.getLogger(__name__)
if not logger.handlers:
    handler = logging.StreamHandler()
    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    handler.setFormatter(formatter)
    logger.addHandler(handler)
    logger.setLevel(logging.INFO)
from typing import Any, Dict, Optional
from openpyxl.utils import get_column_letter
def get_business_column_mapping(workbook: Any, header_row_index: int = 9) -> Dict[str, Dict[str, str]]:
    """Return a mapping of sheet name to column‑letter → header.

    Parameters
    ----------
    workbook: openpyxl.Workbook
        The loaded workbook.
    header_row_index: int, optional
        The row number that contains the column names. Defaults to ``9``.
    """
    mapping: Dict[str, Dict[str, str]] = {}
    for sheet_name in workbook.sheetnames:
        try:
            sheet = workbook[sheet_name]
            header_row = next(
                sheet.iter_rows(min_row=header_row_index,
                                max_row=header_row_index,
                                values_only=True)
            )
        except Exception as e:
            logger.warning(f"Failed to read header row from sheet '{sheet_name}': {e}")
            continue

        col_map: Dict[str, str] = {}
        unnamed_counter = 1
        for idx, header in enumerate(header_row, start=1):
            col_letter = get_column_letter(idx)
            if header is None or (isinstance(header, str) and header.strip() == ""):
                header_str = f"Unnamed{unnamed_counter}"
                unnamed_counter += 1
            else:
                header_str = str(header).strip()
            col_map[col_letter] = header_str
        mapping[sheet_name] = col_map
    return mapping

"""Utilities for extracting business column mappings from Excel worksheets.

The mapping is a dict of sheet name -> {column_letter: header}.
"""

import re
from typing import Dict, Any, Optional
from openpyxl.utils import get_column_letter

# Matches cell references like A1, $B$10, AB123 (strips $ for lookup)
_CELL_REF_RE = re.compile(r'\$?([A-Z]{1,3})\$?\d+')


def get_business_column_mapping(workbook: Any, header_row_index: int = 9) -> Dict[str, Dict[str, str]]:
    """Return a mapping of sheet name to column‑letter → header.

    Parameters
    ----------
    workbook: openpyxl.Workbook
        The loaded workbook.
    header_row_index: int, optional
        The row number that contains the column names. Defaults to ``9``.
    """
    mapping: Dict[str, Dict[str, str]] = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        header_row = next(
            sheet.iter_rows(min_row=header_row_index,
                            max_row=header_row_index,
                            values_only=True)
        )
        col_map: Dict[str, str] = {}
        unnamed_counter = 1
        for idx, header in enumerate(header_row, start=1):
            col_letter = get_column_letter(idx)
            if header is None or (isinstance(header, str) and header.strip() == ""):
                header_str = f"Unnamed{unnamed_counter}"
                unnamed_counter += 1
            else:
                header_str = str(header).strip()
            col_map[col_letter] = header_str
        mapping[sheet_name] = col_map
    return mapping


def resolve_formula_references(
    formula: str,
    col_map: Dict[str, str],
) -> Dict[str, str]:
    """Return {cell_ref: business_name} for every cell reference found in the formula.

    Only references whose column letter exists in col_map are included.
    Example: formula '=SUM(B10:C10)', col_map {'B': 'Revenue', 'C': 'Cost'}
             → {'B10': 'Revenue', 'C10': 'Cost'}
    """
    resolved: Dict[str, str] = {}
    for match in _CELL_REF_RE.finditer(formula):
        col_letter = match.group(1)
        cell_ref = match.group(0).replace('$', '')
        if col_letter in col_map:
            resolved[cell_ref] = col_map[col_letter]
    return resolved


def validate_mappings(
    mapping_v1: Dict[str, Dict[str, str]],
    mapping_v2: Dict[str, Dict[str, str]],
) -> Dict[str, Any]:
    """Compare two workbook mappings and report differences.

    * **added** – columns (headers) that appear only in ``mapping_v2``.
    * **removed** – columns (headers) that disappear in ``mapping_v2``.
    * **moved** – columns whose header is unchanged but the column letter
      shifted (e.g., ``B`` → ``C``). These are *not* reported as added/removed.
    * **changed** – same column letter but different header text.
    """
    sheets_v1 = set(mapping_v1)
    sheets_v2 = set(mapping_v2)

    result: Dict[str, Any] = {
        "added_sheets": sorted(sheets_v2 - sheets_v1),
        "removed_sheets": sorted(sheets_v1 - sheets_v2),
        "sheet_diffs": {},
    }

    # Compare only sheets present in both versions
    for sheet in sheets_v1 & sheets_v2:
        cols_v1 = mapping_v1[sheet]
        cols_v2 = mapping_v2[sheet]

        # Reverse maps: header -> column letter
        header_to_col_v1 = {hdr: col for col, hdr in cols_v1.items()}
        header_to_col_v2 = {hdr: col for col, hdr in cols_v2.items()}

        # Added / removed based on header presence (ignore column position)
        added = {
            header_to_col_v2[h]: h
            for h in set(header_to_col_v2) - set(header_to_col_v1)
        }
        removed = {
            header_to_col_v1[h]: h
            for h in set(header_to_col_v1) - set(header_to_col_v2)
        }

        # Detect moved columns (same header, different column letter)
        moved = {}
        for hdr in set(header_to_col_v1) & set(header_to_col_v2):
            col1 = header_to_col_v1[hdr]
            col2 = header_to_col_v2[hdr]
            if col1 != col2:
                moved[hdr] = {"from": col1, "to": col2}

        # Detect changed header text for the same column letter
        changed = {}
        for col in set(cols_v1) & set(cols_v2):
            hdr1 = cols_v1[col]
            hdr2 = cols_v2[col]
            if hdr1 != hdr2:
                changed[col] = {"from": hdr1, "to": hdr2}

        if added or removed or moved or changed:
            result["sheet_diffs"][sheet] = {
                "added": added,
                "removed": removed,
                "moved": moved,
                "changed": changed,
            }

    return result


def build_mapping_metadata(
    file_name: str,
    version: str,
    mapping: Dict[str, Dict[str, str]],
    formulas: Optional[Dict[str, Dict[str, str]]] = None,
) -> Dict[str, Any]:
    """Build a JSON-ready metadata structure for storage.

    Combines the column mapping with optionally resolved formula references.
    Pass the result directly to persist.save_workbook_record() as the data arg.

    Parameters
    ----------
    mapping:
        Output of get_business_column_mapping().
    formulas:
        Output of formula_extractor.extract_formulas(). When provided, each
        formula cell is enriched with its resolved business-name references.
    """
    metadata: Dict[str, Any] = {
        "file_name": file_name,
        "version": version,
        "column_mappings": mapping,
        "sheets": {},
    }

    for sheet_name, col_map in mapping.items():
        sheet_entry: Dict[str, Any] = {"columns": col_map}

        if formulas and sheet_name in formulas:
            sheet_entry["formulas"] = {
                cell: {
                    "formula": formula,
                    "resolved_references": resolve_formula_references(formula, col_map),
                }
                for cell, formula in formulas[sheet_name].items()
            }

        metadata["sheets"][sheet_name] = sheet_entry

    return metadata
