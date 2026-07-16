# src/excel_exporter.py
"""
Export a comparison report to a formatted Excel audit workbook.

Two sheets are produced:
    Sheet 1 – Summary        Workbook metadata, overall counts, per-sheet breakdown
    Sheet 2 – Detailed Changes  One row per changed/added/removed formula cell

Usage (programmatic):
    from src.excel_exporter import export_report
    export_report(report, "audit_report.xlsx")

The `report` dict is the object returned by compare_service.run_comparison().
"""

from __future__ import annotations

import os
from typing import Any, Dict

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────────────────────
# Palette
# ─────────────────────────────────────────────────────────────────────────────

_C = {
    "header_bg":   "1E3A5F",   # dark navy header rows
    "header_fg":   "FFFFFF",
    "title_bg":    "2B6CB0",   # brighter blue title bar
    "title_fg":    "FFFFFF",
    "section_bg":  "EBF4FF",   # light blue section labels
    "section_fg":  "1E3A5F",
    "added":       "C6EFCE",   # green fill for Added
    "added_fg":    "276221",
    "removed":     "FFC7CE",   # red fill for Removed
    "removed_fg":  "9C0006",
    "logic":       "FFEB9C",   # amber fill for Logic Changed
    "logic_fg":    "9C5700",
    "refshift":    "BDD7EE",   # blue fill for Ref Shift
    "refshift_fg": "1F497D",
    "unchanged":   "F2F2F2",
    "unchanged_fg":"595959",
    "alt_row":     "F7FBFF",   # alternating row tint
    "border":      "D0D7DE",
}


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color="000000", size=10) -> Font:
    return Font(name="Calibri", bold=bold, color=color, size=size)


def _thin_border(sides="all") -> Border:
    thin = Side(style="thin", color=_C["border"])
    none = Side(style=None)
    s = thin if "all" in sides or "l" in sides else none
    e = thin if "all" in sides or "r" in sides else none
    t = thin if "all" in sides or "t" in sides else none
    b = thin if "all" in sides or "b" in sides else none
    return Border(left=s, right=e, top=t, bottom=b)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=False)


def _left_mid() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=False)


def _wrap() -> Alignment:
    return Alignment(horizontal="left", vertical="top", wrap_text=True)

def _clean(value: Any) -> Any:
    """Recursively replace Excel formulas and #REF! errors with empty strings.
    Handles strings that start with '=', as well as '#REF!' patterns.
    """
    if isinstance(value, dict):
        return {k: _clean(v) for k, v in value.items()}
    if isinstance(value, list):
        return [_clean(v) for v in value]
    if isinstance(value, str):
        stripped = value.lstrip()
        if stripped.startswith("=") or stripped.upper().startswith("#REF"):
            return ""
        return value
    return value


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 1 – Summary
# ─────────────────────────────────────────────────────────────────────────────

def _build_summary_sheet(ws, report: Dict[str, Any]) -> None:
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    summary   = report["summary"]
    old_file  = report.get("old_file", "—")
    new_file  = report.get("new_file", "—")
    gen_at    = report.get("generated_at", "—")

    col_widths = [28, 50, 18, 18, 18, 18, 18, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1

    # ── Title bar ────────────────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:H{row}")
    c = ws.cell(row, 1, "Formula Comparison Audit Report")
    c.font      = _font(bold=True, color=_C["title_fg"], size=14)
    c.fill      = _fill(_C["title_bg"])
    c.alignment = _center()
    ws.row_dimensions[row].height = 26
    row += 1

    # ── Metadata block ───────────────────────────────────────────────────────
    meta = [
        ("Old Workbook",      old_file),
        ("New Workbook",      new_file),
        ("Comparison Date",   gen_at),
    ]
    for label, value in meta:
        ws.row_dimensions[row].height = 18
        lc = ws.cell(row, 1, label)
        lc.font = _font(bold=True, color=_C["section_fg"])
        lc.fill = _fill(_C["section_bg"])
        lc.alignment = _left_mid()
        lc.border = _thin_border()

        ws.merge_cells(f"B{row}:H{row}")
        vc = ws.cell(row, 2, value)
        vc.font = _font()
        vc.alignment = _left_mid()
        vc.border = _thin_border()
        row += 1

    row += 1  # blank spacer

    # ── Overall counts section header ────────────────────────────────────────
    ws.merge_cells(f"A{row}:H{row}")
    hc = ws.cell(row, 1, "Overall Change Counts")
    hc.font = _font(bold=True, color=_C["header_fg"])
    hc.fill = _fill(_C["header_bg"])
    hc.alignment = _left_mid()
    ws.row_dimensions[row].height = 20
    row += 1

    overall = [
        ("Added",            summary["total_added"],        _C["added"],    _C["added_fg"]),
        ("Removed",          summary["total_removed"],       _C["removed"],  _C["removed_fg"]),
        ("Logic Changed",    summary["semantic_changes"],    _C["logic"],    _C["logic_fg"]),
        ("Reference Shift",  summary["reference_shifts"],    _C["refshift"], _C["refshift_fg"]),
        ("Unchanged",        summary["total_unchanged"],     _C["unchanged"],_C["unchanged_fg"]),
        ("Total Changes",    summary["total_changes"],       "D9EAD3",       "274E13"),
    ]
    for label, value, bg, fg in overall:
        ws.row_dimensions[row].height = 18
        lc = ws.cell(row, 1, label)
        lc.font = _font(bold=True, color=fg)
        lc.fill = _fill(bg)
        lc.alignment = _left_mid()
        lc.border = _thin_border()

        ws.merge_cells(f"B{row}:H{row}")
        vc = ws.cell(row, 2, value)
        vc.font = _font(bold=True, color=fg)
        vc.fill = _fill(bg)
        vc.alignment = _left_mid()
        vc.border = _thin_border()
        row += 1

    # Sheets added / removed
    if summary["sheets_added"]:
        ws.row_dimensions[row].height = 18
        ws.cell(row, 1, "Sheets Added").font = _font(bold=True, color=_C["added_fg"])
        ws.cell(row, 1).fill = _fill(_C["added"])
        ws.cell(row, 1).alignment = _left_mid()
        ws.merge_cells(f"B{row}:H{row}")
        ws.cell(row, 2, ", ".join(summary["sheets_added"])).font = _font()
        ws.cell(row, 2).alignment = _left_mid()
        row += 1

    if summary["sheets_removed"]:
        ws.row_dimensions[row].height = 18
        ws.cell(row, 1, "Sheets Removed").font = _font(bold=True, color=_C["removed_fg"])
        ws.cell(row, 1).fill = _fill(_C["removed"])
        ws.cell(row, 1).alignment = _left_mid()
        ws.merge_cells(f"B{row}:H{row}")
        ws.cell(row, 2, ", ".join(summary["sheets_removed"])).font = _font()
        ws.cell(row, 2).alignment = _left_mid()
        row += 1

    row += 1  # spacer

    # ── Per-sheet breakdown ───────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:H{row}")
    shc = ws.cell(row, 1, "Per-Sheet Breakdown")
    shc.font = _font(bold=True, color=_C["header_fg"])
    shc.fill = _fill(_C["header_bg"])
    shc.alignment = _left_mid()
    ws.row_dimensions[row].height = 20
    row += 1

    per_sheet_headers = ["Sheet Name", "Added", "Removed", "Logic Changed",
                         "Ref Shift", "Unchanged", "Total Δ"]
    for col_idx, h in enumerate(per_sheet_headers, 1):
        c = ws.cell(row, col_idx, h)
        c.font = _font(bold=True, color=_C["header_fg"])
        c.fill = _fill(_C["header_bg"])
        c.alignment = _center()
        c.border = _thin_border()
    ws.row_dimensions[row].height = 18
    row += 1

    per_sheet = summary.get("per_sheet", {})
    for i, (sheet_name, st) in enumerate(per_sheet.items()):
        bg = _C["alt_row"] if i % 2 else "FFFFFF"
        ws.row_dimensions[row].height = 17
        vals = [
            sheet_name,
            st["added"],
            st["removed"],
            st.get("semantic_changes", 0),
            st.get("reference_shifts", 0),
            st["unchanged"],
            st["added"] + st["removed"] + st["modified"],
        ]
        for col_idx, v in enumerate(vals, 1):
            c = ws.cell(row, col_idx, v)
            c.font = _font()
            c.fill = _fill(bg)
            c.alignment = _left_mid() if col_idx == 1 else _center()
            c.border = _thin_border()
        row += 1

    # Auto-filter on the per-sheet table
    first_data_row = row - len(per_sheet) - 1
    ws.auto_filter.ref = f"A{first_data_row}:G{row - 1}"

    # ── Newly Added Columns ───────────────────────────────────────────────────
    col_mapping  = report.get("column_mapping", {})
    mapping_old  = col_mapping.get("old", {})
    mapping_new  = col_mapping.get("new", {})

    # Collect: columns whose header name appears in new but NOT in old
    new_col_rows = []
    for sname in sorted(mapping_new.keys()):
        old_names = set((mapping_old.get(sname) or {}).values())
        for letter, name in sorted((mapping_new.get(sname) or {}).items()):
            if name and name not in old_names:
                new_col_rows.append((sname, letter, name))

    if new_col_rows:
        row += 1  # spacer
        ws.merge_cells(f"A{row}:H{row}")
        nc = ws.cell(row, 1, "Newly Added Columns")
        nc.font = _font(bold=True, color=_C["header_fg"])
        nc.fill = _fill(_C["header_bg"])
        nc.alignment = _left_mid()
        ws.row_dimensions[row].height = 20
        row += 1

        nc_headers = ["Sheet Name", "Column Letter", "Column Name"]
        for col_idx, h in enumerate(nc_headers, 1):
            c = ws.cell(row, col_idx, h)
            c.font = _font(bold=True, color=_C["header_fg"])
            c.fill = _fill(_C["header_bg"])
            c.alignment = _center()
            c.border = _thin_border()
        ws.row_dimensions[row].height = 18
        row += 1

        for i, (sname, letter, name) in enumerate(new_col_rows):
            bg = _C["alt_row"] if i % 2 else "FFFFFF"
            ws.row_dimensions[row].height = 17
            for col_idx, val in enumerate([sname, letter, name], 1):
                c = ws.cell(row, col_idx, val)
                c.font = _font(bold=(col_idx == 3), color=_C["added_fg"])
                c.fill = _fill(_C["added"])
                c.alignment = _left_mid() if col_idx in (1, 3) else _center()
                c.border = _thin_border()
            row += 1


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 2 – Detailed Changes
# ─────────────────────────────────────────────────────────────────────────────

_CHANGE_TYPE_STYLE = {
    "added":           (_C["added"],    _C["added_fg"],    "Added"),
    "removed":         (_C["removed"],  _C["removed_fg"],  "Removed"),
    "semantic_change": (_C["logic"],    _C["logic_fg"],    "Logic Changed"),
    "reference_shift": (_C["refshift"], _C["refshift_fg"], "Ref Shift"),
}


def _build_detail_sheet(ws, report: Dict[str, Any]) -> None:
    ws.title = "Detailed Changes"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    # Column widths
    widths = [18, 10, 28, 18, 38, 42, 42, 42, 42]
    headers = [
        "Sheet Name", "Cell", "Business Column", "Change Type",
        "Reason", "Old Formula", "New Formula",
        "Old Normalized", "New Normalized",
    ]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
        c = ws.cell(1, i, h)
        c.font      = _font(bold=True, color=_C["header_fg"])
        c.fill      = _fill(_C["header_bg"])
        c.alignment = _center()
        c.border    = _thin_border()
    ws.row_dimensions[1].height = 20

    diff           = report["diff"]
    col_mapping    = report.get("column_mapping", {})
    mapping_old    = col_mapping.get("old", {})
    mapping_new    = col_mapping.get("new", {})

    def _col_letter(cell_addr: str) -> str:
        return "".join(ch for ch in cell_addr if ch.isalpha())

    def _business_name(cell_addr: str, sheet: str, change_kind: str) -> str:
        col = _col_letter(cell_addr)
        mapping = mapping_new if change_kind != "removed" else mapping_old
        return (mapping.get(sheet) or {}).get(col, "")

    row = 2
    for sheet_name, sheet_diff in diff.get("sheet_diffs", {}).items():

        # Added
        added_reasons = sheet_diff.get("added_reasons", {})
        for cell, formula in sorted(sheet_diff.get("added", {}).items()):
            reason_obj = added_reasons.get(cell, {})
            bg, fg, label = _CHANGE_TYPE_STYLE["added"]
            _write_detail_row(
                ws, row,
                sheet_name, cell,
                _business_name(cell, sheet_name, "added"),
                label, fg, bg,
                reason_obj.get("label", ""),
                "",       formula,
                "",       "",
            )
            row += 1

        # Removed
        removed_reasons = sheet_diff.get("removed_reasons", {})
        for cell, formula in sorted(sheet_diff.get("removed", {}).items()):
            reason_obj = removed_reasons.get(cell, {})
            bg, fg, label = _CHANGE_TYPE_STYLE["removed"]
            _write_detail_row(
                ws, row,
                sheet_name, cell,
                _business_name(cell, sheet_name, "removed"),
                label, fg, bg,
                reason_obj.get("label", ""),
                formula,  "",
                "",       "",
            )
            row += 1

        # Modified
        for cell, info in sorted(sheet_diff.get("modified", {}).items()):
            ct  = info["change_type"]          # "semantic_change" | "reference_shift"
            bg, fg, label = _CHANGE_TYPE_STYLE.get(ct, ("FFFFFF", "000000", ct))
            reason_obj = info.get("reason") or {}
            _write_detail_row(
                ws, row,
                sheet_name, cell,
                _business_name(cell, sheet_name, "modified"),
                label, fg, bg,
                reason_obj.get("label", ""),
                info.get("old_formula", ""),
                info.get("new_formula", ""),
                info.get("old_normalized", ""),
                info.get("new_normalized", ""),
            )
            row += 1

    # Auto-filter on the whole table
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _write_detail_row(
    ws, row: int,
    sheet: str, cell: str, biz_col: str,
    change_label: str, fg: str, bg: str,
    reason: str,
    old_f: str, new_f: str,
    old_n: str, new_n: str,
) -> None:
    alt_bg = _C["alt_row"] if row % 2 == 0 else "FFFFFF"

    values = [sheet, cell, biz_col, change_label, reason, old_f, new_f, old_n, new_n]
    for col_idx, val in enumerate(values, 1):
        c = ws.cell(row, col_idx, val)
        c.border    = _thin_border()
        c.alignment = _wrap()

        if col_idx == 4:           # Change Type cell: colored by type
            c.font = _font(bold=True, color=fg)
            c.fill = _fill(bg)
            c.alignment = _center()
        else:
            c.font = _font()
            c.fill = _fill(alt_bg)


# ─────────────────────────────────────────────────────────────────────────────
# Public API
# ─────────────────────────────────────────────────────────────────────────────

def _build_workbook(report: Dict[str, Any]) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws_summary = wb.active
    _build_summary_sheet(ws_summary, report)
    ws_detail = wb.create_sheet()
    _build_detail_sheet(ws_detail, report)
    return wb


def export_report(report: Dict[str, Any], output_path: str) -> str:
    """Generate the Excel audit report and write it to *output_path*.

    Returns the resolved absolute path of the saved file.
    """
    cleaned = _clean(report)
    output_path = os.path.abspath(output_path)
    _build_workbook(cleaned).save(output_path)
    return output_path


def export_report_bytes(report: Dict[str, Any]):
    """Generate the Excel audit report and return it as a BytesIO buffer.

    Useful for streaming the file directly from a web server without
    writing to disk.
    """
    cleaned = _clean(report)
    import io
    buf = io.BytesIO()
    _build_workbook(cleaned).save(buf)
    buf.seek(0)
    return buf
