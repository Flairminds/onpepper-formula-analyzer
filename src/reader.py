# src/reader.py
"""Excel workbook reader using openpyxl.
Provides functions to load a workbook, list sheet names, and extract basic metadata.
"""

import os
from typing import Any, List, Dict
from openpyxl import load_workbook

def load_excel_workbook(file_path: str) -> Any:
    """Load an Excel workbook with formulas preserved.

    Args:
        file_path: Path to the .xlsx file.
    Returns:
        openpyxl Workbook object (data_only=False) to keep formula strings.
    """
    if not os.path.isfile(file_path):
        raise FileNotFoundError(f"Workbook not found: {file_path}")
    # Load without data_only so formulas are retained.
    wb = load_workbook(filename=file_path, data_only=False)
    return wb

def list_sheets(workbook: Any) -> List[str]:
    """Return a list of sheet names in the workbook."""
    return workbook.sheetnames

def get_sheet_metadata(workbook: Any, sheet_name: str, header_row_index: int = 9) -> Dict[str, Any]:
    """Extract basic metadata for a given sheet, using a specific header row.

    Args:
        workbook: The openpyxl Workbook object.
        sheet_name: Name of the sheet to inspect.
        header_row_index: Row number that contains the column headers (default 9).

    Returns:
        Dictionary containing max_row, max_column, and a headers list where empty
        header cells are replaced with unique placeholders (Unnamed1, Unnamed2, …).
    """
    sheet = workbook[sheet_name]
    # Read the designated header row (no fallback – user wants this exact row)
    header_row = next(sheet.iter_rows(min_row=header_row_index,
                                    max_row=header_row_index,
                                    values_only=True))
    # Replace empty/whitespace headers with unique placeholder names
    unnamed_counter = 1
    headers: List[str] = []
    for cell in header_row:
        if cell is None or str(cell).strip() == "":
            headers.append(f"Unnamed{unnamed_counter}")
            unnamed_counter += 1
        else:
            headers.append(str(cell).strip())
    return {
        "max_row": sheet.max_row,
        "max_column": sheet.max_column,
        "headers": headers,
    }

if __name__ == "__main__":
    # Simple demo when run directly
    import sys
    if len(sys.argv) != 2:
        print("Usage: python -m src.reader <path_to_excel>")
        sys.exit(1)
    wb_path = sys.argv[1]
    wb = load_excel_workbook(wb_path)
    print("Sheets:", list_sheets(wb))
    for name in wb.sheetnames:
        print(f"Metadata for sheet {name}:", get_sheet_metadata(wb, name))


