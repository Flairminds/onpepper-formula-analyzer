# test_compare_formulas.py
"""
Simple test script that runs the formula comparison engine on two workbook
versions and prints a concise summary.  It also asserts that the diff
contains at least one changed cell – useful as a sanity‑check in CI.

Usage:
    python test_compare_formulas.py [old.xlsx] [new.xlsx]
If you omit the arguments the script falls back to the two sample workbooks
included in the repository.
"""

import json
import sys
from pathlib import Path

from openpyxl import load_workbook

# Local imports – the src package lives in the repository root.
from src.formula_extractor import extract_formulas
from src.mapping import get_business_column_mapping
from src.formula_normalizer import normalize_all_formulas
from src.formula_comparator import compare_workbooks, generate_summary

# ---------------------------------------------------------------------
# Default workbook pair (the two versions that ship with the project).
# ---------------------------------------------------------------------
DEFAULT_OLD = Path(r"E:\\Excel_Formula_Analyser\\PFLT Sub - Borrowing Base - 07.31.25 ME Truist.xlsx")
DEFAULT_NEW = Path(r"E:\\Excel_Formula_Analyser\\PFLT Sub - Borrowing Base - 08.27.2025 vBorrow_4_TruistSend.xlsx")


def load_workbook_or_fail(path: Path):
    if not path.is_file():
        raise FileNotFoundError(f"Workbook not found: {path}")
    return load_workbook(path, data_only=False)


def main(old_path: Path, new_path: Path) -> None:
    # -----------------------------------------------------------------
    # 1️⃣ Load workbooks
    # -----------------------------------------------------------------
    wb_old = load_workbook_or_fail(old_path)
    wb_new = load_workbook_or_fail(new_path)

    # -----------------------------------------------------------------
    # 2️⃣ Extract raw formulas
    # -----------------------------------------------------------------
    formulas_old = extract_formulas(wb_old)
    formulas_new = extract_formulas(wb_new)

    # -----------------------------------------------------------------
    # 3️⃣ Build column‑header → business‑name mappings (header row = 9)
    # -----------------------------------------------------------------
    mapping_old = get_business_column_mapping(wb_old)
    mapping_new = get_business_column_mapping(wb_new)

    # -----------------------------------------------------------------
    # 4️⃣ Normalise formulas (business‑logic view)
    # -----------------------------------------------------------------
    norm_old = normalize_all_formulas(formulas_old, mapping_old)
    norm_new = normalize_all_formulas(formulas_new, mapping_new)

    # -----------------------------------------------------------------
    # 5️⃣ Run the comparison engine
    # -----------------------------------------------------------------
    diff = compare_workbooks(formulas_old, formulas_new, norm_old, norm_new)
    summary = generate_summary(diff)

    # -----------------------------------------------------------------
    # 6️⃣ Print a readable summary for the user
    # -----------------------------------------------------------------
    print("\n=== FORMULA COMPARISON SUMMARY ===")
    print(json.dumps(summary, indent=2))

    # -----------------------------------------------------------------
    # 7️⃣ Simple sanity‑check – fail fast if nothing changed
    # -----------------------------------------------------------------
    total_changes = summary.get("total_changes", 0)
    if total_changes == 0:
        raise AssertionError(
            "No formula changes detected between the two workbooks. "
            "Either the files are identical or the comparison logic is "
            "incorrect."
        )
    else:
        print(f"\n✅ Detected {total_changes} formula change(s).")

    # -----------------------------------------------------------------
    # 8️⃣ Persist the full diff for later inspection (optional)
    # -----------------------------------------------------------------
    out_file = Path("formula_comparison.json")
    out_file.write_text(json.dumps(diff, indent=2))
    print(f"Full diff written to {out_file.resolve()}")


if __name__ == "__main__":
    # -----------------------------------------------------------------
    # Resolve CLI arguments – two optional paths, otherwise use defaults.
    # -----------------------------------------------------------------
    if len(sys.argv) == 3:
        old_wb = Path(sys.argv[1])
        new_wb = Path(sys.argv[2])
    else:
        old_wb, new_wb = DEFAULT_OLD, DEFAULT_NEW
        print(
            f"Using default workbooks:\n  old → {old_wb.name}\n  new → {new_wb.name}"
        )
    main(old_wb, new_wb)
